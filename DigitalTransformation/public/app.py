from flask import Flask, request, jsonify,render_template
from openpyxl import load_workbook


app = Flask(__name__)
app = Flask(__name__,template_folder="templates")
app = Flask(__name__,static_folder="static")


@app.route("/")
def hello():
    return render_template('index.html')

@app.route("/get_user", methods=['POST'])
def get_user():
    data = request.form.get('data')
    datalist=[]


 
# Give the location of the file
    path = 'C:/Users/Ghost/Desktop/DigitalTransformation/Data/D.xlsx'
    
    # workbook object is created
    wb_obj =load_workbook(path)
    cell_obj=None
    sheet_obj = wb_obj.active                                       
    max_col = sheet_obj.max_column
    for row in sheet_obj.rows:       
        for cell in row:
            # print(cell.value)
            # print(cell.row)
            x=cell.value
            user=x.find(data)
            if user!=-1:
            # if cell.value == "scc.878765654654978.g30":
                for i in range(1, max_col+1):
                    cell_obj = sheet_obj.cell(row = cell.row, column = i)
                    datalist.append(cell_obj.value)
                    print(datalist)
    # Return the selected user as JSON
    return datalist


if __name__ == '__main__':
    app.run("localhost")


# def calories():
#     print("hell")

# @app.route('/', methods=['GET'])
# def calculate_calories():
# #     if request.method == 'POST':
# #         file = request.files['image']
# #         print(file)
# #         fruit_type = request.form['fruit_type']
# #         print(fruit_type)
# #         file_path = "C:/Users/Ghost/Desktop/NN/" + file.filename
# #         file.save(file_path)
#         calories()
#         response = {'calories': 12,'mass':12} 
#         print(response)
#         return jsonify(response)
# if __name__=='__main__':
#     app.run() 
